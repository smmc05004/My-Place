# -*- coding: utf-8 -*-

import json
from os import close
import sys
from openpyxl import load_workbook
import os
from git import Repo


# def i18nToJson2(sheet, config):

#     colHead = []
#     dictLang = {}

#     rowID=""
#     for ri, row in enumerate(sheet.rows):
#         for ci, cell in enumerate(row):
#             #컬럼 헤드
#             if(ri == 0):
#                 colHead.append(cell.value)
#                 dictLang[cell.value] = {}

#             #컬럼 설명 생략
#             elif(ri == 1) :
#                 continue

#             #데이터 
#             else :
#                 if(colHead[ci] == None):
#                     break

#                 dict =  dictLang[colHead[ci]]
#                 if(colHead[ci] == 'ID'):
#                     if(cell.value == None):
#                         break

#                     rowID = int(cell.value)
#                 else :
#                     if(cell.value == None):
#                         dict[rowID] = str(rowID)
#                     else:
#                         dict[rowID] = str(cell.value)
        
#     for code in colHead:
#         print('----------./dist/lang/.json---------', "./dist/lang/{code}.json")
#         jLang = json.dumps(dictLang[code], ensure_ascii=False)
#         stream = open(f"./dist/lang/{code}.json", "w", encoding="utf-8")
#         stream.write(jLang)
#         stream.close()


# def resourceToJson(sheet, info, outPath):

#     colHead = []
#     dictData = {}
 
#     colrums = info["colrums"]
   
#     rowID = None
#     for ri, row in enumerate(sheet.rows):
#         for ci, cell in enumerate(row):
#             #컬럼 헤드
#             if(ri == 0):
#                 if cell.value != None and cell.value not in colrums["ignore"]:
#                     colHead.append(cell.value)
#                 else:
#                     colHead.append(None)
                    
#             #컬럼 설명 생략
#             elif(ri == 1) :
#                 continue

#             #데이터 
#             else :
#                 if(colHead[ci] == None):
#                     continue

#                 if(colHead[ci] == colrums["key"]):
#                     if(cell.value == None):
#                         break

#                     rowID = int(cell.value)
#                     dictData[rowID] = {}
#                 else :
#                     if(rowID != None):
#                         dict = dictData[rowID]
#                         if(cell.value == None):
#                             dict[colHead[ci]] = "#" + str(rowID)
#                         else:
#                             dict[colHead[ci]] = valueToString(cell.value)
                            
#     jLang = json.dumps(dictData, ensure_ascii=False)
#     stream = open(outPath, "w", encoding="utf-8")
#     stream.write(jLang)
#     stream.close()

# def resourceToJsonWithGroup(sheet, info, outPath):

#     colHead = []
#     dictData = {}
 
#     colrums = info["colrums"]
   
#     rowID = None
#     for ri, row in enumerate(sheet.rows):
#         dict = {}
#         for ci, cell in enumerate(row):
#             #컬럼 헤드
#             if(ri == 0):
#                 if cell.value != None and cell.value not in colrums["ignore"]:
#                     colHead.append(cell.value)
#                 else:
#                     colHead.append(None)
                    
#             #컬럼 설명 생략
#             elif(ri == 1) :
#                 continue

#             #데이터 
#             else :
#                 if(colHead[ci] == None):
#                     continue

#                 if(colHead[ci] == colrums["key"]):
#                     if(cell.value == None):
#                         break

#                     rowID = int(cell.value)
#                     try:
#                         dictData[rowID]
#                     except:
#                         dictData[rowID] = []
#                 else :
#                     if(rowID != None):
                        
#                         if(cell.value == None):
#                             dict[colHead[ci]] = "#" + str(rowID)
#                         else:
#                             dict[colHead[ci]] = valueToString(cell.value)

#         if(rowID != None):
#             dictData[rowID].append(dict)
#     jLang = json.dumps(dictData, ensure_ascii=False)
#     stream = open(outPath, "w", encoding="utf-8")
#     stream.write(jLang)
#     stream.close()

# def resourceToJsonWithList(sheet, info, outPath):

#     colHead = []
#     dictData = []
 
#     colrums = info["colrums"]
   
#     for ri, row in enumerate(sheet.rows):
#         dict = {}
#         for ci, cell in enumerate(row):
#             #컬럼 헤드
#             if(ri == 0):
#                 if cell.value != None and cell.value not in colrums["ignore"]:
#                     colHead.append(cell.value)
#                 else:
#                     colHead.append(None)
                    
#             #컬럼 설명 생략
#             elif(ri == 1) :
#                 continue

#             #데이터 
#             else :
#                 if(colHead[ci] == None):
#                     continue
#                 else :
                  
                        
#                     if(cell.value == None):
#                         dict[colHead[ci]] = "#" + str("")
#                     else:
#                         dict[colHead[ci]] = valueToString(cell.value)

#         if(ri >= 2): 
#             dictData.append(dict)
#     jLang = json.dumps(dictData, ensure_ascii=False)
#     stream = open(outPath, "w", encoding="utf-8")
#     stream.write(jLang)
#     stream.close()


# def resourceToConstant(sheet, info, outPath):

#     colHead = []
#     dictData = ""
 
    
   
#     for ri, row in enumerate(sheet.rows):
#         dict = {}
#         for ci, cell in enumerate(row):
#             #컬럼 헤드
#             if(ri == 0):
#                 if cell.value != None:
#                     colHead.append(cell.value)
#                 else:
#                     colHead.append(None)
                    
#             #컬럼 설명 생략
#             elif(ri == 1) :
#                 continue

#             #데이터 
#             else :
#                 if(colHead[ci] == None):
#                     continue
#                 else :
#                     if(cell.value == None):
#                         dict[colHead[ci]] = "#" + str("")
#                     else:
#                         dict[colHead[ci]] = valueToString(cell.value)

#         if(ri >= 2): 
#             dictData += '/* {comment} */ \n'.format(comment = dict["Comment"])
#             dictData += 'export const {name} = {value}; \n'.format(name = dict["Name"], value = valueToConst(dict["Value"]))
#     stream = open(outPath, "w", encoding="utf-8")
#     stream.write(dictData)
#     stream.close()


def valueToString(value):
    if str(type(value)) == "<class 'str'>":
        return value
    elif str(type(value)) == "<class 'float'>" : 
        return str(int(value))
    else :
        return str(value)

def valueToConst(value):
    if str(type(value)) == "<class 'str'>":
        return "'{value}'".format(value = value)
    elif str(type(value)) == "<class 'float'>" : 
        return int(value)
    else :
        return str(value)


def loadExcel(baseUrl, fileName):
    name = '{base}/{name}.xlsx'.format(base = baseUrl, name = fileName)
    return load_workbook(filename = name)

def createDirectory(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print("Error: Failed to create the directory.")

def i18nToJson(sheet, config):
    colHead = []
    dictLang = {}

    key = config["sheets"]["columns"]["key"]

    rowID=""
    for ri, row in enumerate(sheet.rows):
        for ci, cell in enumerate(row):
            #컬럼 헤드
            if(ri == 0):
                colHead.append(cell.value)
                dictLang[cell.value] = {}

            #컬럼 설명 생략
            elif(ri == 1) :
                continue

            #데이터 
            else :
                if colHead[ci] == None:
                    break

                dict = dictLang[colHead[ci]]
                if(colHead[ci] == key):
                    if(cell.value == None):
                        break

                    rowID = str(cell.value)
                else :
                    # 코드 네임과 번호를 분리해서 저장(최적화)
                    codeName = rowID[:3]
                    codeNumber = rowID[3:]

                    if codeName not in dict:
                        dict[codeName] = {}

                    if(cell.value == None):
                        dict[codeName][codeNumber] = str(rowID)
                    else:
                        dict[codeName][codeNumber] = str(cell.value)

    sheetTitle = sheet.title
    for code in colHead:
        if code not in config["sheets"]["columns"]["ignore"]:
            jLang = json.dumps(dictLang[code], ensure_ascii=False)
            createDirectory("./dist/locales/{}".format(sheetTitle))
            stream = open("./dist/locales/{}/{}.json".format(sheetTitle, code), "w", encoding="utf-8")
            stream.write(jLang)
            stream.close()



def buildI18n(config):
    sheetInfo = config["sheets"]

    wb = loadExcel(config['input'], config["name"])
    for sheet in wb.worksheets:    
        if sheet.title not in sheetInfo["ignore"]:
            print('[ Sheet name: {} ]'.format(sheet.title))
            i18nToJson(sheet, config)
    Repo('https://github.com/smmc05004/my-place')
    return True

# 설정파일 읽기
# with open('./configs/manifest.dev.json', 'rt', encoding='utf-8') as f:
#     manifest = json.load(f)

#     # 다국어 처리
#     buildI18n(manifest["locales"])
    
repo = Repo('/test/my-place')
repo.git.add('--all')
print('-------------add---------------')
repo.index.commit('python commit')
print('-------------commit---------------')
origin = repo.remote(name='origin')
origin.push()

# for excelIdx, excel in enumerate(manifest["resource"]):
#     fileName =  excel["name"]
#     name = '{base}/{name}.xlsx'.format(base = manifest["baseUrl"], name = fileName)

#     # 엑셀 파일 불러오기    
#     wb = load_workbook(filename = name)

#     sheetInfo = excel["sheets"]

#     # 시트별 루프
#     for sheet in wb.worksheets:    
#         # 시트 이름 출력
#         print('[ Sheet name: {} ]'.format(sheet.title))

#         for sheetConfigIdx, sheetConfig in enumerate(sheetInfo):   
#             if sheet.title == sheetConfig["name"]:
                
#                 if fileName == "Text":
#                     i18nToJson(sheet, sheetConfig)
#                 elif fileName == "Constants":
#                     outPath = "{base}/constants.ts".format(base = manifest["outDir"])
#                     resourceToConstant(sheet, sheetConfig, outPath)
#                 else :
#                     outPath = "{base}/{name}.json".format(base = manifest["outDir"], name = str.lower(fileName))
#                     if(sheetConfig["type"] == "Group"):
#                         resourceToJsonWithGroup(sheet, sheetConfig, outPath)
#                     elif(sheetConfig["type"] == "List"):
#                         resourceToJsonWithList(sheet, sheetConfig, outPath)
#                     else:
#                         resourceToJson(sheet, sheetConfig, outPath)
                    